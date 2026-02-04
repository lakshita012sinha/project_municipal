from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.utils import timezone
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .models import Survey, SurveyHistory
from .forms import (
    SurveyBasicForm, PropertyOwnerForm, PropertyAddressForm, 
    PropertyDetailsForm, FloorDetailsFormSet, PropertyAmenitiesForm, 
    SurveyGeotaggingForm
)


@login_required
def survey_dashboard(request):
    """
    Survey Dashboard - Main landing page for survey process
    """
    # Check if user has survey permissions
    allowed_roles = ['tax_collector', 'surveyor', 'revenue_inspector', 'commissioner', 'zone_commissioner']
    if request.user.role not in allowed_roles and not request.user.is_staff:
        messages.error(request, 'You do not have permission to access the survey module.')
        return redirect('dashboard')
    
    # Get user's surveys
    user_surveys = Survey.objects.filter(surveyor=request.user).order_by('-created_at')[:10]
    
    # Get survey statistics
    total_surveys = Survey.objects.filter(surveyor=request.user).count()
    completed_surveys = Survey.objects.filter(surveyor=request.user, status='completed').count()
    pending_surveys = Survey.objects.filter(surveyor=request.user, status__in=['draft', 'form_completed', 'geotagged']).count()
    
    context = {
        'user_surveys': user_surveys,
        'total_surveys': total_surveys,
        'completed_surveys': completed_surveys,
        'pending_surveys': pending_surveys,
    }
    
    return render(request, 'survey_process/dashboard.html', context)


@login_required
def create_survey(request):
    """
    Create a new survey - shows options for form filling or geotagging
    """
    allowed_roles = ['tax_collector', 'surveyor', 'revenue_inspector']
    if request.user.role not in allowed_roles and not request.user.is_staff:
        messages.error(request, 'You do not have permission to create surveys.')
        return redirect('survey_process:dashboard')
    
    return render(request, 'survey_process/create_survey.html')


@login_required
def survey_form_entry(request, survey_id=None):
    """
    Survey Form Entry - Complete survey form filling
    """
    survey = None
    if survey_id:
        survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    if request.method == 'POST':
        with transaction.atomic():
            # Basic Survey Form
            basic_form = SurveyBasicForm(request.POST, instance=survey)
            
            if basic_form.is_valid():
                survey = basic_form.save(commit=False)
                if not survey.pk:  # New survey
                    survey.surveyor = request.user
                    survey.status = 'form_completed'
                    # Handle optional fields
                    if not survey.service_no:
                        survey.service_no = None
                    if not survey.unique_key:
                        survey.unique_key = None
                survey.save()
                
                # Create history entry
                SurveyHistory.objects.create(
                    survey=survey,
                    action='form_filled' if survey_id else 'created',
                    performed_by=request.user,
                    notes='Survey form filled'
                )
                
                messages.success(request, 'Survey basic information saved successfully!')
                return redirect('survey_process:survey_owner_details', survey_id=survey.id)
            else:
                messages.error(request, 'Please correct the errors in the form.')
    else:
        basic_form = SurveyBasicForm(instance=survey)
    
    context = {
        'basic_form': basic_form,
        'survey': survey,
        'is_edit': survey is not None,
    }
    
    return render(request, 'survey_process/survey_form_entry.html', context)


@login_required
def survey_owner_details(request, survey_id):
    """
    Property Owner Details Entry
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    try:
        owner = survey.owner
    except:
        owner = None
    
    if request.method == 'POST':
        owner_form = PropertyOwnerForm(request.POST, instance=owner)
        
        if owner_form.is_valid():
            owner = owner_form.save(commit=False)
            owner.survey = survey
            owner.save()
            
            messages.success(request, 'Owner details saved successfully!')
            return redirect('survey_process:survey_address_details', survey_id=survey.id)
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        owner_form = PropertyOwnerForm(instance=owner)
    
    context = {
        'owner_form': owner_form,
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_owner_details.html', context)


@login_required
def survey_address_details(request, survey_id):
    """
    Property Address Details Entry
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    try:
        address = survey.address
    except:
        address = None
    
    if request.method == 'POST':
        address_form = PropertyAddressForm(request.POST, instance=address)
        
        if address_form.is_valid():
            address = address_form.save(commit=False)
            address.survey = survey
            address.save()
            
            messages.success(request, 'Address details saved successfully!')
            return redirect('survey_process:survey_property_details', survey_id=survey.id)
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        address_form = PropertyAddressForm(instance=address)
    
    context = {
        'address_form': address_form,
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_address_details.html', context)


@login_required
def survey_property_details(request, survey_id):
    """
    Property Details Entry
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    try:
        property_details = survey.property
    except:
        property_details = None
    
    if request.method == 'POST':
        property_form = PropertyDetailsForm(request.POST, instance=property_details)
        
        if property_form.is_valid():
            property_details = property_form.save(commit=False)
            property_details.survey = survey
            property_details.save()
            
            messages.success(request, 'Property details saved successfully!')
            return redirect('survey_process:survey_floor_details', survey_id=survey.id)
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        property_form = PropertyDetailsForm(instance=property_details)
    
    context = {
        'property_form': property_form,
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_property_details.html', context)


@login_required
def survey_floor_details(request, survey_id):
    """
    Floor Details Entry
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    if request.method == 'POST':
        formset = FloorDetailsFormSet(request.POST, queryset=survey.floors.all())
        
        if formset.is_valid():
            floors = formset.save(commit=False)
            for floor in floors:
                floor.survey = survey
                floor.save()
            
            # Handle deletions
            for floor in formset.deleted_objects:
                floor.delete()
            
            messages.success(request, 'Floor details saved successfully!')
            return redirect('survey_process:survey_amenities', survey_id=survey.id)
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        formset = FloorDetailsFormSet(queryset=survey.floors.all())
    
    context = {
        'formset': formset,
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_floor_details.html', context)


@login_required
def survey_amenities(request, survey_id):
    """
    Property Amenities Entry
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    try:
        amenities = survey.amenities
    except:
        amenities = None
    
    if request.method == 'POST':
        amenities_form = PropertyAmenitiesForm(request.POST, instance=amenities)
        
        if amenities_form.is_valid():
            amenities = amenities_form.save(commit=False)
            amenities.survey = survey
            amenities.save()
            
            # Update survey status
            survey.status = 'form_completed'
            survey.save()
            
            messages.success(request, 'Survey form completed successfully! Please review all details before saving.')
            return redirect('survey_process:survey_review', survey_id=survey.id)
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        amenities_form = PropertyAmenitiesForm(instance=amenities)
    
    context = {
        'amenities_form': amenities_form,
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_amenities.html', context)


@login_required
def survey_geotagging(request, survey_id):
    """
    Survey Geotagging Entry
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    try:
        geotagging = survey.geotagging
    except:
        geotagging = None
    
    if request.method == 'POST':
        geotagging_form = SurveyGeotaggingForm(request.POST, request.FILES, instance=geotagging)
        
        if geotagging_form.is_valid():
            geotagging = geotagging_form.save(commit=False)
            geotagging.survey = survey
            geotagging.geotagged_by = request.user
            geotagging.save()
            
            # Update survey status
            if survey.status == 'form_completed':
                survey.status = 'completed'
            else:
                survey.status = 'geotagged'
            survey.save()
            
            # Create history entry
            SurveyHistory.objects.create(
                survey=survey,
                action='geotagged',
                performed_by=request.user,
                notes='Survey geotagged with photos'
            )
            
            messages.success(request, 'Survey geotagging completed successfully!')
            return redirect('survey_process:survey_detail', survey_id=survey.id)
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        geotagging_form = SurveyGeotaggingForm(instance=geotagging)
    
    context = {
        'geotagging_form': geotagging_form,
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_geotagging.html', context)


@login_required
def survey_review(request, survey_id):
    """
    Survey Review View - Shows all filled details in non-editable format for review
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    # Check if survey has basic form data
    if not hasattr(survey, 'owner') or not hasattr(survey, 'address'):
        messages.error(request, 'Survey form is incomplete. Please fill all required sections first.')
        return redirect('survey_process:survey_form_edit', survey_id=survey.id)
    
    context = {
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_review.html', context)


@login_required
def survey_save_final(request, survey_id):
    """
    Save survey as final after review
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    if request.method == 'POST':
        # Update survey status to form_completed
        survey.status = 'form_completed'
        survey.save()
        
        # Create history entry
        SurveyHistory.objects.create(
            survey=survey,
            action='completed',
            performed_by=request.user,
            notes='Survey form completed and saved after review'
        )
        
        messages.success(request, 'Survey saved successfully! You can now proceed with geotagging.')
        return redirect('survey_process:survey_detail', survey_id=survey.id)
    
    return redirect('survey_process:survey_review', survey_id=survey.id)


@login_required
def survey_detail(request, survey_id):
    """
    Survey Detail View - Shows complete survey information
    """
    survey = get_object_or_404(Survey, id=survey_id)
    
    # Check permissions
    if survey.surveyor != request.user and request.user.role not in ['commissioner', 'zone_commissioner'] and not request.user.is_staff:
        messages.error(request, 'You do not have permission to view this survey.')
        return redirect('survey_process:dashboard')
    
    context = {
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_detail.html', context)


@login_required
def survey_list(request):
    """
    List all surveys with filtering options
    """
    surveys = Survey.objects.all().order_by('-created_at')
    
    # Filter by surveyor if not admin
    if request.user.role not in ['commissioner', 'zone_commissioner'] and not request.user.is_staff:
        surveys = surveys.filter(surveyor=request.user)
    
    # Apply filters
    status_filter = request.GET.get('status')
    if status_filter:
        surveys = surveys.filter(status=status_filter)
    
    surveyor_filter = request.GET.get('surveyor')
    if surveyor_filter:
        surveys = surveys.filter(surveyor_id=surveyor_filter)
    
    search_query = request.GET.get('search')
    if search_query:
        surveys = surveys.filter(
            models.Q(service_no__icontains=search_query) |
            models.Q(unique_key__icontains=search_query) |
            models.Q(survey_number__icontains=search_query)
        )
    
    context = {
        'surveys': surveys,
        'status_filter': status_filter,
        'surveyor_filter': surveyor_filter,
        'search_query': search_query,
    }
    
    return render(request, 'survey_process/survey_list.html', context)


@login_required
def get_location(request):
    """
    AJAX endpoint to get current location using HTML5 Geolocation
    """
    return JsonResponse({'status': 'success'})


@login_required
def delete_survey(request, survey_id):
    """
    Delete survey (only for draft surveys)
    """
    survey = get_object_or_404(Survey, id=survey_id, surveyor=request.user)
    
    if survey.status != 'draft':
        messages.error(request, 'Only draft surveys can be deleted.')
        return redirect('survey_process:survey_detail', survey_id=survey.id)
    
    if request.method == 'POST':
        survey.delete()
        messages.success(request, 'Survey deleted successfully.')
        return redirect('survey_process:dashboard')
    
    context = {
        'survey': survey,
    }
    
    return render(request, 'survey_process/survey_delete.html', context)


@login_required
def export_survey_data(request):
    """
    Export survey data to Excel with date filtering
    """
    # Check if Excel functionality is available
    if not EXCEL_AVAILABLE:
        messages.error(request, 'Excel export functionality is not available. Please install openpyxl.')
        return redirect('survey_process:dashboard')
    
    # Check permissions
    allowed_roles = ['commissioner', 'zone_commissioner', 'revenue_inspector', 'back_office_head', 'MIS']
    if request.user.role not in allowed_roles and not request.user.is_staff:
        messages.error(request, 'You do not have permission to export survey data.')
        return redirect('survey_process:dashboard')
    
    if request.method == 'POST':
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        zone = request.POST.get('zone', 'ALL')
        ward_no = request.POST.get('ward_no', 'ALL')
        is_taxable = request.POST.get('is_taxable', 'ALL')
        is_assessed = request.POST.get('is_assessed', 'ALL')
        operator_name = request.POST.get('operator_name', 'ALL')
        survey_no = request.POST.get('survey_no', '')
        service_no = request.POST.get('service_no', '')
        
        # Parse dates
        try:
            if date_from:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            if date_to:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format. Please use YYYY-MM-DD format.')
            return render(request, 'survey_process/export_form.html')
        
        # Filter surveys
        surveys = Survey.objects.select_related('owner', 'address', 'property', 'amenities', 'geotagging', 'surveyor').all()
        
        if date_from:
            surveys = surveys.filter(created_at__date__gte=date_from)
        if date_to:
            surveys = surveys.filter(created_at__date__lte=date_to)
        if service_no:
            surveys = surveys.filter(service_no__icontains=service_no)
        if operator_name != 'ALL':
            surveys = surveys.filter(surveyor__emp_name__icontains=operator_name)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Survey Report"
        
        # Define headers based on the image format
        headers = [
            '#', 'Survey No.', 'Service No.', 'Zone', 'Ward No.', 'Property Usage',
            'Property Category', 'DLC Ward Area', 'Plot Area', 'Plinth Area',
            'Total Builtup Area', 'Road Type', 'Owner Name', 'Gender', 'Mobile No.',
            'House No.', 'Mohalla/Colony', 'Building Name', 'Road Name', 'Sector',
            'Pincode', 'Latitude', 'Longitude', 'Surveyor', 'Survey Date', 'Status'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        row_num = 2
        for idx, survey in enumerate(surveys, 1):
            try:
                owner = survey.owner
            except:
                owner = None
            
            try:
                address = survey.address
            except:
                address = None
            
            try:
                property_details = survey.property
            except:
                property_details = None
            
            try:
                geotagging = survey.geotagging
            except:
                geotagging = None
            
            # Calculate total built-up area
            total_builtup_area = sum([floor.built_up_area_sq_ft for floor in survey.floors.all()])
            
            row_data = [
                idx,  # #
                survey.unique_key,  # Survey No.
                survey.service_no,  # Service No.
                property_details.area_name if property_details else '',  # Zone
                property_details.ward_no if property_details else '',  # Ward No.
                property_details.get_use_of_property_display() if property_details else '',  # Property Usage
                'Residential',  # Property Category (default)
                '',  # DLC Ward Area (not in current model)
                f"{property_details.plot_area_sq_yd}" if property_details else '',  # Plot Area
                f"{property_details.plinth_area_sq_yd}" if property_details else '',  # Plinth Area
                f"{total_builtup_area}",  # Total Builtup Area
                property_details.get_road_type_display() if property_details else '',  # Road Type
                owner.owner_name if owner else '',  # Owner Name
                owner.get_gender_display() if owner else '',  # Gender
                owner.mobile_no if owner else '',  # Mobile No.
                address.house_no if address else '',  # House No.
                address.mohalla_colony_name if address else '',  # Mohalla/Colony
                address.building_name if address else '',  # Building Name
                address.road_name if address else '',  # Road Name
                address.sector if address else '',  # Sector
                address.pincode if address else '',  # Pincode
                f"{geotagging.latitude}" if geotagging else '',  # Latitude
                f"{geotagging.longitude}" if geotagging else '',  # Longitude
                survey.surveyor.emp_name if survey.surveyor.emp_name else survey.surveyor.username,  # Surveyor
                survey.created_at.strftime('%d-%m-%Y'),  # Survey Date
                survey.get_status_display(),  # Status
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename with date range
        filename = f"Survey_Report_{date_from or 'all'}_{date_to or 'all'}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    # GET request - show export form
    context = {
        'today': date.today().strftime('%Y-%m-%d'),
    }
    return render(request, 'survey_process/export_form.html', context)